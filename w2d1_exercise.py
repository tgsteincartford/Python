# ITP Week 2 Day 1 Exercise

# SCENARIO: You're a store owner receiving the inventory report for this month.
# You will receive the product order spreadsheet soon and it is easier to calculate your order
# if your inventory was also on a spreadsheet! Recreate the following spreadsheet with Python: 

# don't forget your appropriate imports.
from pathlib import Path
from openpyxl import Workbook

BASE_DIR = Path(__file__).resolve().parent  # week_2/day_1
SPREADSHEETS_DIR = BASE_DIR.parent / "spreadsheets"
SPREADSHEETS_DIR.mkdir(exist_ok=True)

file_path = SPREADSHEETS_DIR / "inventory.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "CURRENT_MONTH_INVENTORY"

# ... append headers + data ...

wb.save(file_path)
print("Saved to:", file_path)

#1) assign the title of the initial active sheet to "CURRENT_MONTH_INVENTORY"
#2 Headers
headers = ["product_name", "product_id", "max_amount", "reorder_threshold", "quantity"]
ws.append(headers)

#3 Data (strings for product_id)
inventory = [
    ["oreo", "2323", 1000, 300, 743],
    ["coke", "6545", 500, 100, 101],
    ["pepsi", "3456", 200, 50, 137],
    ["lays_chip", "4567", 1500, 500, 364],
    ["pringles", "2134", 2000, 600, 120],
    ["sour_worms", "2362", 100, 10, 85],
    ["choco_cookies", "0923", 200, 25, 24],
    ["donuts", "2786", 200, 25, 12],
    ["hot_dogs", "6723", 100, 10, 39],
    ["ice_cream", "9237", 200, 50, 234],
    ["gum", "2092", 3500, 1000, 1232],
    ["pretzels", "8246", 100, 5, 11],
    ["kit_kat", "9276", 1000, 250, 249],
]
#   product_name    product_id  max_amount      reorder_threshold   quantity
# 	oreo            2323        1000            300                 743
# 	coke            6545        500             100                 101
# 	pepsi	        3456        200             50                  137
# 	lays_chip	    4567        1500            500                 364
# 	pringles	    2134        2000            600                 120
# 	sour_worms	    2362        100             10                  85
# 	choco_cookies   0923	    200             25                  24
# 	donuts	        2786        200             25                  12
# 	hot_dogs	    6723        100             10                  39
# 	ice_cream	    9237        200             50                  234
# 	gum	            2092        3500            1000                1232
# 	pretzels        8246	    100             5                   11
# 	kit_kat	        9276        1000            250                 249

# TIP: create a list of each column (ie. product_names = [oreo, ...]) and use those to loop through :)

#4 Write Data rows
for row in inventory:
    ws.append(row)

#5 - save your file
wb.save("./spreadsheets/inventory.xlsx")

#6 - double checks
print(ws.title)
print(ws.max_row, ws.max_column) #14 rows (1 header + 13 items) + 5 columns
for r in ws.iter_rows(min_row=1, max_row=4, values_only=True):
    print(r)
