# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)


# EASY MODE

# import the appropriate modules (you have 3)
import requests
import json
from openpyxl import Workbook

# character_url = "https://rickandmortyapi.com/api/character"
# set up a workbook and worksheet titled "Rick and Morty Characters"
wb = Workbook()
ws = wb.active
ws.title = "Rick and Morty Characters"
character_url = "https://rickandmortyapi.com/api/character"

# # assign a variable 'data' with the returned GET request
data = requests.get(character_url).json()
# create the appropriate headers in openpyxl for all of the keys for a single character
from openpyxl.utils import get_column_letter
# loop through all of the 'results' of the data to populate the rows and columns for each character
headers = ["id", "name", "status", "species", "type", "gender", "origin", "location", "image", "episode", "url", "created"]
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    ws[f"{col_letter}1"] = header
row_num = 2
for character in data['results']:
    ws[f"A{row_num}"] = character['id']
    ws[f"B{row_num}"] = character['name']
    ws[f"C{row_num}"] = character['status']
    ws[f"D{row_num}"] = character['species']
    ws[f"E{row_num}"] = character['type']
    ws[f"F{row_num}"] = character['gender']
    ws[f"G{row_num}"] = character['origin']['name']
    ws[f"H{row_num}"] = character['location']['name']
    ws[f"I{row_num}"] = character['image']
    ws[f"J{row_num}"] = ", ".join(character['episode'])
    ws[f"K{row_num}"] = character['url']
    ws[f"L{row_num}"] = character['created']
    row_num += 1
# NOTE: due to the headers, the rows need to be offset by one!

# MEDIUM MODE

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"
ws_location = wb.create_sheet(title="Rick and Morty Locations")
ws_episode = wb.create_sheet(title="Rick and Morty Episodes")
# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
# "https://rickandmortyapi.com/api/location"
# "https://rickandmortyapi.com/api/episode"
location_url = "https://rickandmortyapi.com/api/location"
episode_url = "https://rickandmortyapi.com/api/episode"
# make the appropriate GET requests to the two new urls and assign them to variables location_data and
# populate the new worksheets appropriately with all of the data!
location_data = requests.get(location_url).json()
episode_data = requests.get(episode_url).json()
# populate locations worksheet
location_headers = ["id", "name", "type", "dimension", "residents", "url", "created"]
for col_num, header in enumerate(location_headers, 1):
    col_letter = get_column_letter(col_num)
    ws_location[f"{col_letter}1"] = header
row_num = 2
for location in location_data['results']:
    ws_location[f"A{row_num}"] = location['id']
    ws_location[f"B{row_num}"] = location['name']
    ws_location[f"C{row_num}"] = location['type']
    ws_location[f"D{row_num}"] = location['dimension']
    ws_location[f"E{row_num}"] = ", ".join(location['residents'])
    ws_location[f"F{row_num}"] = location['url']
    ws_location[f"G{row_num}"] = location['created']
    row_num += 1
# populate episodes worksheet
episode_headers = ["id", "name", "air_date", "episode", "characters", "url", "created"]
for col_num, header in enumerate(episode_headers, 1):
    col_letter = get_column_letter(col_num)
    ws_episode[f"{col_letter}1"] = header
row_num = 2
for episode in episode_data['results']:
    ws_episode[f"A{row_num}"] = episode['id']
    ws_episode[f"B{row_num}"] = episode['name']
    ws_episode[f"C{row_num}"] = episode['air_date']
    ws_episode[f"D{row_num}"] = episode['episode']
    ws_episode[f"E{row_num}"] = ", ".join(episode['characters'])
    ws_episode[f"F{row_num}"] = episode['url']
    ws_episode[f"G{row_num}"] = episode['created']
    row_num += 1
# NOTE: don't forget your headers!

# HARD MODE
# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)
while data['info']['next']:
    next_url = data['info']['next']
    next_data = requests.get(next_url).json()
    data['results'].extend(next_data['results'])
    data['info'] = next_data['info']
# NIGHTMARE
# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url
# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)
#^^^uhhh no thanks^^^## I did ok already :D

wb.save("./spreadsheets/exercise.xlsx")
