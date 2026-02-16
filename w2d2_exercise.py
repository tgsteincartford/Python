# ITP Week 2 Day 2 Exercise

# import the appropriate method from the correct module

# 1) Import the workbook that you updated in today's practice from
# "./spreadsheets/inventory.xlsx"
# #2) access and store the appropriate worksheet to the variable 'ws'
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent          # .../week_2/day_2
file_path = BASE_DIR / "spreadsheets" / "inventory.xlsx"  # .../week_2/day_2/spreadsheets/inventory.xlsx

print("Loading:", file_path)
print("Exists?", file_path.exists())

wb = load_workbook(file_path)
ws = wb.active


# 3)Define a function called add_order_amount that takes in a single parameter called 'row'
    # IF the quantity is less or equal to than the threshold, then calculate the order_amount (max_amount - quantity) and assign the value to that row, column 6
# TIP: create variables for quantity, threshold, max_amount that retrieves the values first for cleanliness

def add_order_amount(row):
    quantity = ws.cell(row=row, column=5).value
    threshold = ws.cell(row=row, column=4).value
    max_amount = ws.cell(row=row, column=3).value

    if quantity <= threshold:
        ws.cell(row=row, column=6).value = max_amount - quantity

for row in range(2, ws.max_row + 1):
    add_order_amount(row)

# 4) Perform a for..in loop through the range(2, len(inventory.rows)) #<-- this is wrong, no inventory variable
#   - call the function add_order_amount() passing in the number of the range

#5) save the worksheet
wb.save("spreadsheets/inventory.xlsx")


