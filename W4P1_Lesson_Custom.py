import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Fetch cities data
cities_url = "https://cityweatherclass.free.beeceptor.com/cities"
cities_response = requests.get(cities_url)
cities_text = cities_response.text
cities_data = json.loads(cities_text)

#check record
print("First city record:", cities_data[0])

# Fetch weather data
weather_url = "https://cityweatherclass.free.beeceptor.com/weather"
weather_response = requests.get(weather_url)
weather_text = weather_response.text
weather_data = json.loads(weather_text)

# check record
print("First weather record:", weather_data[0])

# weather lookup dictionary
weather_dict = {item['city']: item for item in weather_data}

#merge the data
merged_data = []
for city in cities_data:
    city_name = city["city"]
    if city_name in weather_dict:
        merged_item = {
            'City': city_name,
            'Population': city['population'],
            'Region': city['region'],
            'Temperature': weather_dict[city_name]['temperature'],
            'Humidity': weather_dict[city_name]['humidity']
        }
        merged_data.append(merged_item)

#print first record to check it
print("First merged record:", merged_data[0])

#Write to excel
wb = Workbook()
ws = wb.active
ws.title = "City Weather Report"

#Headers
headers = ['city', 'population', 'region', 'temperature (°F)', 'Humidity (%)']
for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

#Data rows
for row_num, item in enumerate(merged_data, start=2):
    ws.cell(row=row_num, column=1, value=item['City'])
    ws.cell(row=row_num, column=2, value=item['Population'])
    ws.cell(row=row_num, column=3, value=item['Region'])
    ws.cell(row=row_num, column=4, value=item['Temperature'])
    ws.cell(row=row_num, column=5, value=item['Humidity'])

# auto-adjust widths (A trick I looked up...:P)
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

#Save it before you brave it
wb.save('city_weather_report.xlsx')
print("Jeddah destroyed, Admiral")
